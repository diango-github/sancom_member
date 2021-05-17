from django.urls import reverse_lazy
from django.views import generic
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import (
     get_user_model, logout as auth_logout,
)
from .forms import UserCreateForm
from .models import User, File
import openpyxl

User = get_user_model()


class Top(generic.TemplateView):
    params = {
        'visiter':''
    }

    def get(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            id = str(self.request.user.id)
            user_name= self.request.user.get_username()
            if not File.objects.filter(owner=self.request.user).exists():
                file = File()
                file.owner = User.objects.filter(email=user_name).first()
                file.filename1 = 'SancomContents_' + id + '.xlsx'
                file.filename2 = 'voa_' + id + '.xlsx'
                file.filename3 = 'librivox_' + id + '.xlsx'

                wb=openpyxl.load_workbook("./static/SancomContents.xlsx")
                #wb=openpyxl.load_workbook("./static/voa.xlsx")
                #wb=openpyxl.load_workbook("./static/librivox.xlsx")
                wb.save("./static/" + file.filename1)
                #wb.save("./static/" + file.filename2)
                #wb.save("./static/" + file.filename3)
                file.save()
            self.params['visiter'] = user_name
        return render(self.request,'index.html', self.params) # top から index へ変更


#class Top(generic.TemplateView):
    #template_name = 'top.html'


class SignUpView(generic.CreateView):
    form_class = UserCreateForm
    success_url = reverse_lazy('login')
    template_name = 'registration/signup.html'


class ProfileView(LoginRequiredMixin, generic.View):

    def get(self, *args, **kwargs):
        return render(self.request,'registration/profile.html')


class DeleteView(LoginRequiredMixin, generic.View):

    def get(self, *args, **kwargs):
        user = User.objects.get(email=self.request.user.email)
        user.is_active = False
        user.save()
        auth_logout(self.request)
        return render(self.request,'registration/delete_complete.html')
